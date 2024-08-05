import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def get_google_sheet_as_dataframe(url):
    df = pd.read_csv(url)
    return df


def get_amazon_data():
    return get_google_sheet_as_dataframe('https://docs.google.com/spreadsheets/d/1rC7PEHzLF5k3Oc73fJ3_7tAZXySp0oJ30lX3DHJTxmE/export?gid=0&format=csv').to_dict(orient='records')


# def get_flipcart_data():
#     return get_google_sheet_as_dataframe('https://docs.google.com/spreadsheets/d/1kgqFc2pYijdyYhcDB0JJ3dwnY5ONj5zMlb576hdaZj0/export?gid=31823053&format=csv').to_dict(orient='records')


# def get_1mg_data():
#     return get_google_sheet_as_dataframe('https://docs.google.com/spreadsheets/d/1kgqFc2pYijdyYhcDB0JJ3dwnY5ONj5zMlb576hdaZj0/export?gid=1811929831&format=csv').to_dict(orient='records')


# def get_nykaa_data():
#     return get_google_sheet_as_dataframe('https://docs.google.com/spreadsheets/d/1kgqFc2pYijdyYhcDB0JJ3dwnY5ONj5zMlb576hdaZj0/export?gid=1704332346&format=csv').to_dict(orient='records')


# def get_hyugalife_data():
#     return get_google_sheet_as_dataframe('https://docs.google.com/spreadsheets/d/1kgqFc2pYijdyYhcDB0JJ3dwnY5ONj5zMlb576hdaZj0/export?gid=1022515349&format=csv').to_dict(orient='records')


# def color_cells(df):
#     def color_cell(s):
#         if s.name.startswith('scraped_'):
#             source_col = 'source_' + s.name[8:]
#             style_list = []
#             for scraped, source in zip(s, df[source_col]):
#                 try:
#                     if scraped > source:
#                         style_list.append('background-color: green')
#                     elif scraped < source:
#                         style_list.append('background-color: red')
#                     else:
#                         style_list.append('background-color: gray')
#                 except:
#                     style_list.append('')
#             return style_list
#         else:
#             return [''] * len(s)

#     return df.style.apply(color_cell)


# def compile_data(amazon_data, flipcart_data, one_mg_data, nykaa_data, hugalife_data):
#     excel_writer = pd.ExcelWriter('data/output.xlsx', engine='xlsxwriter')
    
#     color_cells(pd.DataFrame(data=amazon_data)).to_excel(excel_writer, sheet_name='Amazon', index=False)
#     color_cells(pd.DataFrame(data=flipcart_data)).to_excel(excel_writer, sheet_name='Flipcart', index=False)
#     color_cells(pd.DataFrame(data=one_mg_data)).to_excel(excel_writer, sheet_name='1mg', index=False)
#     color_cells(pd.DataFrame(data=nykaa_data)).to_excel(excel_writer, sheet_name='Nykaa', index=False)
#     color_cells(pd.DataFrame(data=hugalife_data)).to_excel(excel_writer, sheet_name='HyugaLife', index=False)

#     excel_writer._save()


def save_data(data_list):

    df = pd.DataFrame(data_list)

    # Create a Pandas Excel writer using openpyxl as the engine
    with pd.ExcelWriter('data/output.xlsx', engine='openpyxl') as writer:
        # Convert the dataframe to an Excel object
        df.to_excel(writer, index=False, sheet_name='Sheet1')

        # Access the openpyxl workbook and sheet objects
        workbook  = writer.book
        worksheet = writer.sheets['Sheet1']

        # Define a light gray fill
        fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        # Apply the fill to the specific columns (3rd, 5th, 7th, 9th, and 11th)
        for col in [3, 5, 7, 9, 11]:
            for row in range(2, len(df) + 2):  # +2 to account for header and 1-based index in openpyxl
                cell = worksheet.cell(row=row, column=col)
                cell.fill = fill

        # Save the workbook
        workbook.save('data/output.xlsx')
